import openpyxl
import xlsxwriter
import time
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

TD = time.localtime()
Ph = "./spreadsheets/" + str(TD.tm_mday) + "-" + str(TD.tm_mon) + "-" + str(TD.tm_year) + ".xlsx"
TD = str(TD.tm_mday) + "-" + str(TD.tm_mon) + "-" + str(TD.tm_year)


def Get(var):
    if var == "temp":
        var = 2
    if var == "humid":
        var = 3
    Data = []
    wb = openpyxl.load_workbook(Ph)
    ws = wb.active
    MaxRow = ws.max_row
    for i in range(2, MaxRow+1):
        D = ws.cell(row=(i), column=(var))
        Data.append(float(D.value))
    return Data


def GetTime():
    Time = []
    wb = openpyxl.load_workbook(Ph)
    ws = wb.active
    MaxRow = ws.max_row
    for i in range(2, MaxRow+1):
        T = ws.cell(row=(i), column=1)
        Time.append(T.value)
    Time = [datetime.strptime(d, "%H:%M") for d in Time]
    return Time


def init():
    workbook = xlsxwriter.Workbook(Ph)
    worksheet = workbook.add_worksheet()
    worksheet.write(0, 0, 'Thời gian')
    worksheet.write(0, 1, 'Nhiệt độ')
    worksheet.write(0, 2, 'Độ ẩm')
    workbook.close()


def write(Time, Temp, Humid):
    wb = openpyxl.load_workbook(Ph)
    ws = wb.active
    MaxRow = ws.max_row
    MaxColumn = ws.max_column
    workbook = xlsxwriter.Workbook(Ph)
    worksheet = workbook.add_worksheet()
    # copy the old values over
    for i in range(1, MaxColumn+1):
        for n in range(1, MaxRow+1):
            name = ws.cell(row=(n), column=(i))
            if i == 1:
                worksheet.write((n-1), (i-1), str((name.value)))
            else:
                worksheet.write((n-1), (i-1), (name.value))
    worksheet.write((MaxRow), 0, Time)
    worksheet.write((MaxRow), 1, Temp)
    worksheet.write((MaxRow), 2, Humid)
    workbook.close()


def CurrentTime():
    Time = time.localtime()
    Time = str(Time.tm_hour) + ":" + str(Time.tm_min)
    return Time


def Graph():
    fig, ax = plt.subplots(figsize=(8.8, 4), constrained_layout=True)
    plt.plot(GetTime(), Get("humid"), label="Độ ẩm")
    plt.plot(GetTime(), Get("temp"), label="Nhiệt độ")
    plt.ylabel("Nhiệt độ và độ ẩm")
    plt.xlabel("Thời gian")
    plt.title("Biểu đồ thể hiện nhiệt độ và độ ẩm không khí qua thời gian ngày " + TD )
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    plt.legend()
    plt.savefig('graph.png')
    plt.savefig("./spreadsheets/" + TD + ".png")


def echo(file, text):
    file = open(file+".txt", "w")
    file.write(str(text))
    file.close
